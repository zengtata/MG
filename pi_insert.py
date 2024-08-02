import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

def process_files(input_files, output_file_path):
    # Initialize lists for collecting all data
    all_no = []
    all_model_name = []
    all_model = []
    all_material_code = []
    all_interior_color = []
    all_exterior_color = []
    all_qty = []
    all_unit_price = []
    all_price = []
    all_pi_number = []
    all_mc_pi = []
    all_date_value = []
    all_importer = []
    all_exporter = []
    all_total_payment = []
    all_tt_value = []
    all_lc_value = []

    for input_file_path in input_files:
        # Initialize per-file variables
        no = []
        model_name = []
        model = []
        material_code = []
        interior_color = []
        exterior_color = []
        qty = []
        unit_price = []
        price = []
        pi_number = ""
        mc_pi = []
        date_value = ""
        importer = ""
        exporter = ""
        total_payment = 0
        tt_value = 0
        lc_value = 0

        # Load the input workbook and select the active sheet
        wb_input = openpyxl.load_workbook(input_file_path)
        sheet_input = wb_input.active

        # Parsing the rows
        for row in sheet_input.iter_rows(values_only=True):
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    if cell == "Date:":
                        date_value = row[idx + 1]
                        if isinstance(date_value, datetime):
                            date_value = date_value.strftime('%Y-%m-%d')
                    elif cell.startswith("Invoice Number:"):
                        parts = cell.split(":")
                        if len(parts) > 1:
                            pi_number = parts[1].strip()
                    elif "T/T" in cell:
                        parts = cell.split(" ")
                        if len(parts) > 1:
                            tt = parts[0].strip()

                    elif "L/C" in cell:
                        parts = cell.split(" ")
                        if len(parts) > 1:
                            lc = parts[0].strip()


        # Parsing the columns
        for col in sheet_input.iter_cols(values_only=True):
            for idx, cell in enumerate(col):
                if cell == "No.":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            no.append(col[next_idx])
                        if next_idx + 1 < len(col) and col[next_idx + 1] == "TOTAL Qty:":
                            break
                elif cell == "Model":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            model.append(col[next_idx])
                            parts = col[next_idx].split(" ")
                            model_name.append(parts[0])
                        if next_idx + 1 < len(col) and col[next_idx + 1] == "TOTAL Qty:":
                            break
                elif cell == "Importer:":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            importer = col[next_idx]
                            break
                elif cell == "Exporter:":
                    exporter = col[idx + 1]
                elif cell == "Material Code":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            material_code.append(col[next_idx])
                            mc_pi.append(col[next_idx] + pi_number)
                        else:
                            break
                elif cell == "Unit Price":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            unit_price.append(col[next_idx])
                        else:
                            break
                elif cell == "Interior \nColor":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            interior_color.append(col[next_idx])
                        else:
                            break
                elif cell == "Exterior\n Color":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            exterior_color.append(col[next_idx])
                        else:
                            break
                elif cell == "Qty":
                    for next_idx in range(idx + 1, idx + len(no) + 1):
                        if next_idx < len(col) and col[next_idx] is not None:
                            qty.append(col[next_idx])
                        else:
                            break

        for i in range(len(unit_price)):
            price.append(unit_price[i] * qty[i])

        for i in range(len(price)):
            total_payment += price[i]

        lc_value = int(lc.strip('%')) / 100 * float(total_payment)
        tt_value = int(tt.strip('%')) / 100 * float(total_payment)

        # Append per-file data to the main lists
        all_no.extend(no)
        all_model_name.extend(model_name)
        all_model.extend(model)
        all_material_code.extend(material_code)
        all_interior_color.extend(interior_color)
        all_exterior_color.extend(exterior_color)
        all_qty.extend(qty)
        all_unit_price.extend(unit_price)
        all_price.extend(price)
        all_pi_number.extend([pi_number] * len(no))
        all_mc_pi.extend(mc_pi)
        all_date_value.extend([date_value] * len(no))
        all_importer.extend([importer] * len(no))
        all_exporter.extend([exporter] * len(no))
        all_total_payment.extend([total_payment] * len(no))
        all_tt_value.extend([tt_value] * len(no))
        all_lc_value.extend([lc_value] * len(no))

    # Load or create the output workbook and select the specified sheet
    if os.path.exists(output_file_path):
        wb_output = openpyxl.load_workbook(output_file_path)
        if "PI_YTD" in wb_output.sheetnames:
            sheet_output = wb_output["PI_YTD"]
        else:
            sheet_output = wb_output.create_sheet("PI_YTD")
    else:
        wb_output = Workbook()
        sheet_output = wb_output.active
        sheet_output.title = "PI_YTD"

    # Prepare the data to be inserted (keeping column 3 (Trim level) empty)
    data = list(
        zip(all_no, all_model_name, [''] * len(all_no), all_model, all_material_code, all_interior_color, all_exterior_color, all_qty, all_unit_price, all_price,
            all_pi_number, all_mc_pi, all_date_value, all_importer, all_exporter, all_total_payment, all_tt_value, all_lc_value))

    # Retrieve existing mc_pi values from the sheet
    existing_mc_pi = set()
    mc_pi_column_index = 12  # Assuming "MC & PI" is the 12th column (1-based index)

    for row in sheet_output.iter_rows(min_row=2, max_row=sheet_output.max_row, min_col=mc_pi_column_index,
                                      max_col=mc_pi_column_index, values_only=True):
        if row[0] is not None:
            existing_mc_pi.add(row[0])

    # Insert the data into the output sheet
    for row in data:
        mc_pi_value = row[11]
        if mc_pi_value not in existing_mc_pi:
            sheet_output.append(row)
            existing_mc_pi.add(mc_pi_value)
        else:
            print(f"Skipped row with MC PI {mc_pi_value}")

    # Save the output workbook
    wb_output.save(output_file_path)
    messagebox.showinfo("Success", f"Data has been processed and saved to {output_file_path}")

# Create the main window
root = tk.Tk()
root.title("PI Excel Data Extractor")
root.geometry("400x400")

# Frame for the file list with scrollbar
frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

# Create a scrollbar
scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a listbox to display the selected files
file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

# Configure scrollbar
scrollbar.config(command=file_listbox.yview)

def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        # Clear the listbox before adding new files
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)

def save_file():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialfile="PI_extracted_data.xlsx")
    if save_path:
        files = file_listbox.get(0, tk.END)
        if files:
            process_files(files, save_path)
        else:
            messagebox.showwarning("No Files", "Please select at least one file to process.")

# Buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_file)
save_as_button.pack(side=tk.LEFT, padx=10)

# Run the GUI event loop
root.mainloop()
