import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import os


def extract_data(sheet):
    vin = []
    engineNumber = []
    material_code = []
    disposition = []
    color = []
    licenseName = []
    trackingNumber = []
    shipmentPlanNumber = []
    salesOrderNumber = []
    destinationPort = []
    vesselName = []
    pi = []

    headers = {
        "发运单号": trackingNumber,
        "发运计划编号": shipmentPlanNumber,
        "销售订单号": salesOrderNumber,
        "配置": disposition,
        "颜色": color,
        "实车": vin,
        "物料编码": material_code,
        "发动机号": engineNumber,
        "许可证名称": licenseName,
        "目的港": destinationPort,
        "船名航次": vesselName,
        "PI": pi
    }

    for col in sheet.iter_cols(values_only=True):
        header = col[0]
        if header in headers:
            headers[header].extend(col[1:])

    return {
        'trackingNumber': trackingNumber,
        'shipmentPlanNumber': shipmentPlanNumber,
        'salesOrderNumber': salesOrderNumber,
        'disposition': disposition,
        'color': color,
        'vin': vin,
        'material_code': material_code,
        'engineNumber': engineNumber,
        'licenseName': licenseName,
        'destinationPort': destinationPort,
        'vesselName': vesselName,
        'pi': pi
    }


def process_files(input_files, output_filename):
    if os.path.exists(output_filename):
        output_workbook = openpyxl.load_workbook(output_filename)
        # Remove default sheets if they exist
        default_sheets = [sheet for sheet in output_workbook.sheetnames if sheet != "DN"]
        for sheet in default_sheets:
            del output_workbook[sheet]

        if "DN" in output_workbook.sheetnames:
            new_sheet = output_workbook["DN"]
            # Track existing data (including the source file column)
            existing_data = {tuple(row) for row in new_sheet.iter_rows(min_row=2, values_only=True)}
            # Track existing source files
            existing_files = {row[-1] for row in existing_data}
        else:
            new_sheet = output_workbook.create_sheet(title="DN")
            titles = ["VIN", "Engine Number", "Material Code", "Disposition", "Color", "License Name",
                      "Tracking Number", "Shipment Plan Number", "Sales Order Number", "PI", "Destination Port",
                      "Vessel Name", "Source File"]
            for col_idx, title in enumerate(titles, start=1):
                new_sheet.cell(row=1, column=col_idx, value=title)
            existing_data = set()
            existing_files = set()
    else:
        output_workbook = Workbook()
        new_sheet = output_workbook.create_sheet(title="DN")
        titles = ["VIN", "Engine Number", "Material Code", "Disposition", "Color", "License Name",
                  "Tracking Number", "Shipment Plan Number", "Sales Order Number", "PI", "Destination Port",
                  "Vessel Name", "Source File"]
        for col_idx, title in enumerate(titles, start=1):
            new_sheet.cell(row=1, column=col_idx, value=title)
        existing_data = set()
        existing_files = set()

    row_idx = new_sheet.max_row + 1

    for input_file in input_files:
        try:
            if input_file in existing_files:
                continue

            workbook = openpyxl.load_workbook(input_file, data_only=True)
            sheets_to_process = ["Main", "VehicleInformation", "车辆清单"]

            for sheet_name in sheets_to_process:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    data = extract_data(sheet)

                    for i in range(len(data["vin"])):
                        row_data = [
                            data["vin"][i] if i < len(data["vin"]) else '',
                            data["engineNumber"][i] if i < len(data["engineNumber"]) else '',
                            data["material_code"][i] if i < len(data["material_code"]) else '',
                            data["disposition"][i] if i < len(data["disposition"]) else '',
                            data["color"][i] if i < len(data["color"]) else '',
                            data["licenseName"][i] if i < len(data["licenseName"]) else '',
                            data["trackingNumber"][i] if i < len(data["trackingNumber"]) else '',
                            data["shipmentPlanNumber"][i] if i < len(data["shipmentPlanNumber"]) else '',
                            data["salesOrderNumber"][i] if i < len(data["salesOrderNumber"]) else '',
                            data["pi"][i] if i < len(data["pi"]) else '',
                            data["destinationPort"][i] if i < len(data["destinationPort"]) else '',
                            data["vesselName"][i] if i < len(data["vesselName"]) else '',
                            input_file
                        ]

                        # Create a tuple of the entire row data including the "Source File" column
                        row_data_key = tuple(row_data)

                        # Check if this entry already exists
                        if row_data_key not in existing_data:
                            # Write the extracted data to the main sheet
                            for col_idx, value in enumerate(row_data, start=1):
                                new_sheet.cell(row=row_idx, column=col_idx, value=value)
                            existing_data.add(row_data_key)
                            row_idx += 1

            existing_files.add(input_file)

        except Exception as e:
            print(f"Failed to process {input_file}: {e}")

    output_workbook.save(output_filename)
    messagebox.showinfo("Process Complete",
                        f"Data has been saved to {output_filename}\n"
                        f"Number of rows created: {row_idx - 2}")

# GUI setup
def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)


def save_file():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialfile="DN_extracted_data.xlsx")
    if save_path:
        files = file_listbox.get(0, tk.END)
        process_files(files, save_path)


root = tk.Tk()
root.title("DN Excel Data Extractor")
root.geometry("600x400")  # Increased size for better UX

frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

scrollbar.config(command=file_listbox.yview)

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_file)
save_as_button.pack(side=tk.LEFT, padx=10)

root.mainloop()
