import openpyxl
from openpyxl import Workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# Global variables for data extraction
invoiceNO = ""
date_value = ""
importer = ""
total_quantity = ""
total_payment = ""
tt = ""
lc = ""
tt_value = ""
lc_value = ""
currency = ""
port_of_unload = ""
material_code = []
model_quantity = []
unit_price = []


def extract_data(sheet1):
    global invoiceNO, date_value, importer, total_quantity, total_payment, tt, lc, tt_value, lc_value, currency, port_of_unload, material_code, model_quantity, unit_price

    # Reset lists for each new file processed
    material_code = []
    model_quantity = []
    unit_price = []

    for row in sheet1.iter_rows(values_only=True):
        for idx, cell in enumerate(row):
            if isinstance(cell, str):
                if cell == "Date:":
                    date_value = row[idx + 1]
                    if isinstance(date_value, datetime):
                        date_value = date_value.strftime('%Y-%m-%d')
                elif cell.startswith("Invoice Number:"):
                    parts = cell.split(":")
                    if len(parts) > 1:
                        invoiceNO = parts[1].strip()
                elif cell == "Port of Unloading:":
                    port_of_unload = row[idx + 2]
                elif cell == "TOTAL Qty:":
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is not None:
                            total_quantity = row[next_idx]
                            break
                elif cell == "TOTAL PAYMENT :":
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is not None:
                            total_payment = row[next_idx]
                            break
                elif "T/T" in cell:
                    parts = cell.split(" ")
                    if len(parts) > 1:
                        tt = parts[0].strip()
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is not None:
                            tt_value = row[next_idx]
                            break
                elif "L/C" in cell:
                    parts = cell.split(" ")
                    if len(parts) > 1:
                        lc = parts[0].strip()
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is not None:
                            lc_value = row[next_idx]
                            break
                elif cell == "Currency：":
                    currency = row[idx + 1]

    for col in sheet1.iter_cols(values_only=True):
        for idx, cell in enumerate(col):
            if isinstance(cell, str):
                if cell == "Importer:":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            importer = col[next_idx]
                            break
                elif cell == "Material Code":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            material_code.append(col[next_idx])
                        else:
                            break
                elif cell == "Qty":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            model_quantity.append(col[next_idx])
                        else:
                            break
                elif cell == "Unit Price":
                    for next_idx in range(idx + 1, len(col)):
                        if col[next_idx] is not None:
                            unit_price.append(col[next_idx])
                        else:
                            break

    return {
        "Invoice_NO.": invoiceNO,
        "Date": date_value,
        "Importer": importer,
        "Total Quantity": total_quantity,
        "Total Payment": total_payment,
        "T/T": tt,
        "L/C": lc,
        "T/T Value": tt_value,
        "L/C Value": lc_value,
        "Currency": currency,
        "Port of Unloading": port_of_unload,
        "Material Code": material_code,
        "Model Quantity": model_quantity,
        "Unit Price": unit_price
    }


def data_exists_in_sheet(sheet, data, detailed=False):
    if detailed:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                if (data["Material Code"] and data["Model Quantity"] and data["Unit Price"] and
                        row[3] == data["Invoice_NO."] and
                        row[0] in data["Material Code"] and
                        row[1] in data["Model Quantity"] and
                        row[2] in data["Unit Price"]):
                    return True
            except TypeError:
                continue
    else:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                if (row[0] == data["Invoice_NO."] and
                        row[1] == data["Date"] and
                        row[2] == data["Importer"] and
                        row[3] == data["Port of Unloading"] and
                        row[4] == data["Total Quantity"] and
                        row[5] == data["Total Payment"] and
                        row[6] == data["T/T"] and
                        row[7] == data["T/T Value"] and
                        row[8] == data["L/C"] and
                        row[9] == data["L/C Value"] and
                        row[10] == data["Currency"]):
                    return True
            except TypeError:
                continue
    return False


def process_files(input_files, output_filename):
    if os.path.exists(output_filename):
        # Load the existing workbook
        output_workbook = openpyxl.load_workbook(output_filename)
        # Get the existing sheet or create a new one
        if "PI_extracted_data" in output_workbook.sheetnames:
            new_sheet = output_workbook["PI_extracted_data"]
        else:
            new_sheet = output_workbook.create_sheet(title="PI_extracted_data")
            # Write column titles
            titles = ["Invoice_NO.", "Date", "Importer", "Port of Unloading", "Total Quantity", "Total Payment",
                      "T/T", "T/T Value", "L/C", "L/C Value", "Currency"]
            for col_idx, title in enumerate(titles, start=1):
                new_sheet.cell(row=1, column=col_idx, value=title)
    else:
        # Create a new workbook for output
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)  # Remove the default sheet created
        # Create a single sheet for all data
        new_sheet = output_workbook.create_sheet(title="PI_extracted_data")
        # Write column titles
        titles = ["Invoice_NO.", "Date", "Importer", "Port of Unloading", "Total Quantity", "Total Payment",
                  "T/T", "T/T Value", "L/C", "L/C Value", "Currency"]
        for col_idx, title in enumerate(titles, start=1):
            new_sheet.cell(row=1, column=col_idx, value=title)

    # Create or get the detailed data sheet
    if "detailed_data" in output_workbook.sheetnames:
        detailed_sheet = output_workbook["detailed_data"]
    else:
        detailed_sheet = output_workbook.create_sheet(title="detailed_data")
        # Write column titles for detailed data
        detailed_titles = ["Material Code", "Model Quantity", "Unit Price", "Invoice_NO."]
        for col_idx, title in enumerate(detailed_titles, start=1):
            detailed_sheet.cell(row=1, column=col_idx, value=title)

    # Create sets to keep track of existing data
    existing_invoice_nos = set()
    existing_detailed_data = set()

    # Populate existing invoice numbers and detailed data
    for row in new_sheet.iter_rows(min_row=2, values_only=True):
        existing_invoice_nos.add(row[0])

    for row in detailed_sheet.iter_rows(min_row=2, values_only=True):
        if row[3] and (row[0], row[1], row[2], row[3]) not in existing_detailed_data:
            existing_detailed_data.add((row[0], row[1], row[2], row[3]))

    # Find the next available row in the existing or new sheets
    row_idx = new_sheet.max_row + 1
    detailed_row_idx = detailed_sheet.max_row + 1

    for input_file in input_files:
        try:
            # Load the input workbook with data_only=True to get cell values instead of formulas
            workbook = openpyxl.load_workbook(input_file, data_only=True)

            # Extract data from sheet1
            sheet1 = workbook['PI']
            data = extract_data(sheet1)

            # Check if the main data already exists
            if data["Invoice_NO."] not in existing_invoice_nos:
                # Write the extracted data to the main sheet
                new_sheet.cell(row=row_idx, column=1, value=data["Invoice_NO."])
                new_sheet.cell(row=row_idx, column=2, value=data["Date"])
                new_sheet.cell(row=row_idx, column=3, value=data["Importer"])
                new_sheet.cell(row=row_idx, column=4, value=data["Port of Unloading"])
                new_sheet.cell(row=row_idx, column=5, value=data["Total Quantity"])
                new_sheet.cell(row=row_idx, column=6, value=data["Total Payment"])
                new_sheet.cell(row=row_idx, column=7, value=data["T/T"])
                new_sheet.cell(row=row_idx, column=8, value=data["T/T Value"])
                new_sheet.cell(row=row_idx, column=9, value=data["L/C"])
                new_sheet.cell(row=row_idx, column=10, value=data["L/C Value"])
                new_sheet.cell(row=row_idx, column=11, value=data["Currency"])
                existing_invoice_nos.add(data["Invoice_NO."])
                row_idx += 1  # Move to the next row for the next set of data

            # Handle detailed data separately
            for i in range(len(data["Material Code"])):
                detailed_data = (
                    data["Material Code"][i],
                    data["Model Quantity"][i],
                    data["Unit Price"][i],
                    data["Invoice_NO."]
                )
                if detailed_data not in existing_detailed_data:
                    detailed_sheet.cell(row=detailed_row_idx, column=1, value=data["Material Code"][i])
                    detailed_sheet.cell(row=detailed_row_idx, column=2, value=data["Model Quantity"][i])
                    detailed_sheet.cell(row=detailed_row_idx, column=3, value=data["Unit Price"][i])
                    detailed_sheet.cell(row=detailed_row_idx, column=4, value=data["Invoice_NO."])
                    existing_detailed_data.add(detailed_data)
                    detailed_row_idx += 1

        except Exception as e:
            print(f"Failed to process {input_file}: {e}")
            messagebox.showwarning("File Error", f"Failed to process {input_file}: {e}")

    # Save the workbook
    output_workbook.save(output_filename)

    # Show a message box indicating completion
    messagebox.showinfo("Process Complete",
                        f"Data has been saved to {output_filename}\n"
                        f"Number of rows created in main sheet: {row_idx - 2}\n"
                        f"Number of rows created in detailed sheet: {detailed_row_idx - 2}")


# Create the main window
root = tk.Tk()
root.title("PI Excel Data Extractor")
root.geometry("400x400")

# Frame for the file list with scrollbar
frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

# Create a scrollbar
scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a text widget to display the selected files
file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

# Configure scrollbar
scrollbar.config(command=file_listbox.yview)

def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        # Clear the listbox before adding new files
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)

def save_file():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialfile="PI_extracted_data.xlsx")
    if save_path:
        files = file_listbox.get(0, tk.END)
        process_files(files, save_path)

# Buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_file)
save_as_button.pack(side=tk.LEFT, padx=10)

# Run the GUI event loop
root.mainloop()

