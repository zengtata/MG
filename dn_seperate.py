import openpyxl
from openpyxl import Workbook
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def process_file(file_path):
    # Load the input workbook and select the sheet named 'Main'
    wb_input = openpyxl.load_workbook(file_path)
    sheet_input = wb_input['Main']

    # Read headers from the first row
    headers = [cell.value for cell in sheet_input[1]]
    header_to_index = {header: idx for idx, header in enumerate(headers)}

    # Check if 'Dest.' column exists
    if 'Dest.' not in header_to_index:
        raise ValueError("The column 'Dest.' was not found in the input sheet.")

    dest_column_index = header_to_index['Dest.']

    # Collect data by destination
    data_by_dest = {}

    for row in sheet_input.iter_rows(min_row=2, values_only=True):
        dest = row[dest_column_index]
        if dest not in data_by_dest:
            data_by_dest[dest] = []
        data_by_dest[dest].append(row)

    # Create or load the output workbook
    if os.path.exists(file_path):
        wb_output = openpyxl.load_workbook(file_path)
    else:
        wb_output = Workbook()

    # Create new sheets for each destination if they do not already exist
    for dest, rows in data_by_dest.items():
        sheet_name = dest[:31]  # Excel sheet names have a max length of 31 characters
        if sheet_name not in wb_output.sheetnames:
            sheet_output = wb_output.create_sheet(sheet_name)
        else:
            sheet_output = wb_output[sheet_name]

        # Write header row
        for col, header in enumerate(headers, start=1):
            sheet_output.cell(row=1, column=col, value=header)

        # Write data rows
        for row_index, row_data in enumerate(rows, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet_output.cell(row=row_index, column=col_index, value=cell_value)

    # Save the output workbook
    wb_output.save(file_path)
    messagebox.showinfo("Success", f"Data has been separated into sheets based on the 'Dest.' column and saved to {file_path}.")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)

def run_processing():
    file_path = entry_file.get()
    if not file_path:
        messagebox.showwarning("Input Required", "Please specify a file.")
        return
    try:
        process_file(file_path)
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Create the main window
root = tk.Tk()
root.title("Excel Data Separator")
root.geometry("400x150")

# Frame for file selection
frame_file = tk.Frame(root)
frame_file.pack(pady=10, padx=10, fill='x')

label_file = tk.Label(frame_file, text="Select File:")
label_file.pack(side=tk.LEFT)
entry_file = tk.Entry(frame_file, width=40)
entry_file.pack(side=tk.LEFT, padx=5)
button_browse_file = tk.Button(frame_file, text="Browse", command=browse_file)
button_browse_file.pack(side=tk.LEFT)

# Button to start processing
button_process = tk.Button(root, text="Process", command=run_processing)
button_process.pack(pady=20)

# Run the GUI event loop
root.mainloop()
