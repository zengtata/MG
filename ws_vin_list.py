import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import openpyxl
import os


def process_files(input_files, output_file):
    # Assuming input_files contains a list of file paths
    for input_file_path in input_files:
        # Load the input workbook and select the sheet named 'Main'
        wb_input = openpyxl.load_workbook(input_file_path)
        sheet_input = wb_input['Main']

        vin = []
        material_code = []
        specification = []
        destinationCountry = []
        vesselName = []
        distributor = []
        pi = []
        model = []
        tracker = []

        headers = {
            "实车": vin,
            "物料编码": material_code,
            "许可证名称": specification,
            "Dest.": destinationCountry,
            "船名航次": vesselName,
            "PI": pi
        }

        # Collect data based on headers
        for col in sheet_input.iter_cols(values_only=True):
            header = col[0]
            if header in headers:
                headers[header].extend(col[1:])

        for i in range(len(specification)):
            parts = specification[i].split(" ")
            model.append(parts[0])

            if destinationCountry[i] == "HU":
                distributor.append("Duna Motors Disztribúció Kft.")
            elif destinationCountry[i] == "CZ":
                distributor.append("AB Motor")
            elif destinationCountry[i] == "GR":
                distributor.append("Aiglon S.A")
            elif destinationCountry[i] == "HR":
                distributor.append("GRAND AUTOMOTIVE ADRIATIC d.o.o.")
            elif destinationCountry[i] == "RO":
                distributor.append("Quantum Auto Max srl")
            elif destinationCountry[i] == "PL":
                distributor.append("KROTOSKI")
            elif destinationCountry[i] == "SK":
                distributor.append("AB Motor")

            tracker.append(material_code[i] + pi[i] + vesselName[i])

        # Load or create the output workbook and select the specified sheet
        if os.path.exists(output_file):
            wb_output = openpyxl.load_workbook(output_file)
            if "VIN LIST" in wb_output.sheetnames:
                sheet_output = wb_output["VIN LIST"]
            else:
                sheet_output = wb_output.create_sheet("VIN LIST")
        else:
            wb_output = Workbook()
            sheet_output = wb_output.active
            sheet_output.title = "VIN LIST"

        # Retrieve existing VIN values from the output sheet
        existing_vin = set()
        vin_index = 2  # Assuming VIN is the 2nd column (1-based index)

        for row in sheet_output.iter_rows(min_row=2, max_row=sheet_output.max_row, min_col=vin_index, max_col=vin_index,
                                          values_only=True):
            existing_vin.add(row[0])

        # Prepare the data to be inserted (keeping specified columns empty)
        data = list(zip(
            [''] * len(vin),  # Column 1: Keep empty
            vin,  # Column 2: VIN
            material_code,  # Column 3: Material code
            specification,  # Column 4: Specification
            destinationCountry,  # Column 5: Destination country
            vesselName,  # Column 6: Vessel
            distributor,  # Column 7: Distributor
            [''] * len(vin),  # Column 8: ATD
            pi,  # Column 9: PI number
            [''] * len(vin),  # Column 10: Sales Contract number
            [''] * len(vin),  # Column 11: Invoice number
            model,  # Column 12: Model
            [''] * len(vin),  # Column 13: CIF Price
            [''] * len(vin),  # Column 14: CIF price Q3
            [''] * len(vin),  # Column 15: Variance
            [''] * len(vin),  # Column 16: Trim
            tracker  # Column 17: Tracker
        ))

        # Insert the data into the output sheet
        inserted_count = 0
        for row in data:
            vin_value = row[1]  # VIN is at index 1 in the data list
            if vin_value not in existing_vin:
                sheet_output.append(row)
                existing_vin.add(vin_value)  # Add the newly inserted VIN to the set
                inserted_count += 1

        # Save the output workbook
        wb_output.save(output_file)

        # Notify the user
        messagebox.showinfo("Processing Complete",
                            f"Processed {len(input_files)} files. {inserted_count} rows inserted.")


# Create the main window
root = tk.Tk()
root.title("WS_VIN_Data Extractor")
root.geometry("400x400")

# Frame for the file list with scrollbar
frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

# Create a scrollbar
scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a listbox to display the selected files
file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

# Configure scrollbar
scrollbar.config(command=file_listbox.yview)


def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        # Clear the listbox before adding new files
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)


def save_file():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialfile="ws_vin_data.xlsx")
    if save_path:
        files = file_listbox.get(0, tk.END)
        if files:
            process_files(files, save_path)
        else:
            messagebox.showwarning("No Files", "Please select at least one file to process.")


# Buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_file)
save_as_button.pack(side=tk.LEFT, padx=10)

# Run the GUI event loop
root.mainloop()
