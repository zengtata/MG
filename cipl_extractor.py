import openpyxl
from openpyxl import Workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os

invoiceNO = ""
date_value = ""
sale_contractNO = ""
seller = ""
to_value = ""
delivery_term = ""
total_unit = ""
total_value = ""
delivery_number = ""
vin_numbers = []


def extract_data(sheet1):
    global invoiceNO, date_value, sale_contractNO, seller, to_value, delivery_term, total_unit, total_value, delivery_number

    for row in sheet1.iter_rows(values_only=True):
        for idx, cell in enumerate(row):
            if cell == "INVOICE NO.:":
                invoiceNO = row[idx + 1]
            elif cell == "DATE:":
                date_value = row[idx + 1]
                if isinstance(date_value, datetime):
                    date_value = date_value.strftime('%Y-%m-%d')
            elif cell == "SALE CONTRACT NO.:":
                sale_contractNO = row[idx + 1]
            elif cell == "SELLER: ":
                seller = []
                for next_idx in range(idx + 1, len(row)):
                    if row[next_idx] is None:
                        break
                    seller.append(row[next_idx])
                seller = " ".join(seller)
            elif cell == "TO:":
                to_values = []
                for next_idx in range(idx + 1, len(row)):
                    if row[next_idx] is None:
                        break
                    to_values.append(row[next_idx])
                to_value = " ".join(to_values)
            elif cell == "DELIVERY TERM:":
                delivery_term = row[idx + 1]
            elif cell == "TOTAL":
                total_unit = row[idx + 3]
            elif cell is None and idx + 2 < len(row) and row[idx + 1] == "EUR" and row[idx - 1] is None:
                total_value = row[idx + 2]
            elif isinstance(cell, str) and cell.startswith("DELIVERY NO.:"):
                parts = cell.split(":")
                if len(parts) > 1:
                    delivery_number = parts[1].strip()

    return {
        "Invoice_NO.": invoiceNO,
        "Date": date_value,
        "Sale_contract_NO": sale_contractNO,
        "Seller": seller,
        "To": to_value,
        "Delivery_term": delivery_term,
        "Invoice_total_unit": total_unit,
        "Invoice_total_value": total_value,
        "Delivery_number": delivery_number
    }


def extract_vin_numbers(sheet2):
    vin_numbers = set()  # Use a set to avoid duplicates

    for col in sheet2.iter_cols(values_only=True):
        for idx, cell in enumerate(col):
            if isinstance(cell, str) and cell.startswith("LS") and len(cell) == 17:
                vin_numbers.add(cell)  # Add to set to ensure uniqueness

    vin_numbers = list(vin_numbers)  # Convert back to list for further processing

    return vin_numbers


def process_files(input_files, output_filename):
    if os.path.exists(output_filename):
        # Load the existing workbook
        output_workbook = openpyxl.load_workbook(output_filename)
        # Get the existing sheet or create a new one
        if "CIPL_extracted_data" in output_workbook.sheetnames:
            new_sheet = output_workbook["CIPL_extracted_data"]
        else:
            new_sheet = output_workbook.create_sheet(title="CIPL_extracted_data")
            # Write column titles
            titles = ["VIN Numbers", "Invoice_NO.", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                      "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]
            for col_idx, title in enumerate(titles, start=1):
                new_sheet.cell(row=1, column=col_idx, value=title)
    else:
        # Create a new workbook for output
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)  # Remove the default sheet created
        # Create a single sheet for all data
        new_sheet = output_workbook.create_sheet(title="CIPL_extracted_data")
        # Write column titles
        titles = ["VIN Numbers", "Invoice_NO.", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                  "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]
        for col_idx, title in enumerate(titles, start=1):
            new_sheet.cell(row=1, column=col_idx, value=title)

    row_idx = new_sheet.max_row + 1

    for input_file in input_files:
        try:
            # Load the input workbook with data_only=True to get cell values instead of formulas
            workbook = openpyxl.load_workbook(input_file, data_only=True)

            # Extract data from sheet1
            sheet1 = workbook['CI']
            data = extract_data(sheet1)

            # Extract VIN numbers from sheet2
            sheet2 = workbook['PL']
            vin_numbers = extract_vin_numbers(sheet2)

            # Write VIN numbers and extracted data to the new sheet
            for vin in vin_numbers:
                # Check if VIN already exists in the sheet
                data_exists = False
                for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row, values_only=True):
                    if row[0] == vin and row[1] == data["Invoice_NO."] and row[2] == data["Date"] and \
                            row[3] == data["Sale_contract_NO"] and row[4] == data["Seller"] and \
                            row[5] == data["To"] and row[6] == data["Delivery_term"] and \
                            row[7] == data["Invoice_total_unit"] and row[8] == data["Invoice_total_value"] and \
                            row[9] == data["Delivery_number"]:
                        data_exists = True
                        break

                if not data_exists:
                    # Write VIN numbers and extracted data to the new sheet
                    new_sheet.cell(row=row_idx, column=1, value=vin)
                    col_idx = 2
                    for key in ["Invoice_NO.", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                                "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]:
                        new_sheet.cell(row=row_idx, column=col_idx, value=data[key])
                        col_idx += 1
                    row_idx += 1

        except Exception as e:
            print(f"Failed to process {input_file}: {e}")
            messagebox.showwarning("File Error", f"Failed to process {input_file}: {e}")

    # Save the new workbook with the user-specified filename
    output_workbook.save(output_filename)

    # Show a message box indicating completion
    messagebox.showinfo("Process Complete",
                        f"Data has been saved to {output_filename}\n"
                        f"Number of rows created: {row_idx - 2}")


# GUI Setup

# Create the main window
root = tk.Tk()
root.title("CIPL Excel Data Extractor")
root.geometry("400x400")

# Frame for the file list with scrollbar
frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

# Create a scrollbar
scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a text widget to display the selected files
file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

# Configure scrollbar
scrollbar.config(command=file_listbox.yview)

def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        # Clear the listbox before adding new files
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)

def save_file():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialfile="CIPL_extracted_data.xlsx")
    if save_path:
        files = file_listbox.get(0, tk.END)
        process_files(files, save_path)

# Buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_file)
save_as_button.pack(side=tk.LEFT, padx=10)

# Run the GUI event loop
root.mainloop()
