import os
import tkinter as tk
from tkinter import filedialog, messagebox
import xlrd
from openpyxl import Workbook


# Function to convert .xls to .xlsx
def convert_xls_to_xlsx(files, save_dir):
    for file in files:
        xls_path = file
        xlsx_path = os.path.join(save_dir, os.path.basename(file).replace('.xls', '.xlsx'))

        xls_workbook = xlrd.open_workbook(xls_path)

        xlsx_workbook = Workbook()
        xlsx_workbook.remove(xlsx_workbook.active)  # Remove the default sheet created

        for sheet_index in range(xls_workbook.nsheets):
            xls_sheet = xls_workbook.sheet_by_index(sheet_index)
            xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)

            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

        xlsx_workbook.save(xlsx_path)
        print(f"Converted {file} to {xlsx_path}")


# Function to browse and select files
def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xls")])
    if files:
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)


# Function to select save directory and convert files
def save_files():
    save_dir = filedialog.askdirectory()
    if save_dir:
        files = file_listbox.get(0, tk.END)
        if files:
            convert_xls_to_xlsx(files, save_dir)
            messagebox.showinfo("Success", "Files converted successfully!")
        else:
            messagebox.showwarning("Warning", "No files selected!")


# Create the main window
root = tk.Tk()
root.title("Excel Converter")
root.geometry("400x400")

# Frame for the file list with scrollbar
frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

# Create a scrollbar
scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a listbox to display the selected files
file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

# Configure scrollbar
scrollbar.config(command=file_listbox.yview)

# Buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_as_button = tk.Button(button_frame, text="Save As", command=save_files)
save_as_button.pack(side=tk.LEFT, padx=10)

# Run the GUI event loop
root.mainloop()
